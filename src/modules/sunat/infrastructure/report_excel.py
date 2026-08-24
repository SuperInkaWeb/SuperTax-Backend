"""Genera el reporte Excel de un job de descarga SUNAT a partir de sus resultados."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.modules.sunat.infrastructure.comprobante_xml import TIPOS

_HEADERS = [
    "Comprobante", "Tipo", "Emisor", "RUC", "PDF", "XML",
    "Estado", "Monto", "Fecha", "Descripción",
]
_ANCHOS = [16, 14, 32, 12, 6, 6, 12, 14, 12, 55]


def generar_reporte(resultados: list[dict]) -> bytes:
    """Devuelve el .xlsx (bytes) con una fila por comprobante."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte SUNAT"

    ws.append(_HEADERS)
    relleno = PatternFill("solid", fgColor="FF0E4F4A")
    for col in range(1, len(_HEADERS) + 1):
        celda = ws.cell(row=1, column=col)
        celda.font = Font(bold=True, color="FFFFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(vertical="center", wrap_text=True)

    for r in resultados:
        tipo = str(r.get("tipo", ""))
        monto = f"{r.get('moneda', '')} {r.get('monto', '')}".strip()
        ws.append([
            r.get("id", ""),
            TIPOS.get(tipo, tipo),
            r.get("emisor", ""),
            r.get("ruc", ""),
            "Sí" if r.get("pdf") else "No",
            "Sí" if r.get("xml") else "No",
            r.get("estado", ""),
            monto,
            r.get("fecha", ""),
            r.get("descripcion", ""),
        ])

    for i, ancho in enumerate(_ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    for fila in ws.iter_rows(min_row=2):
        fila[9].alignment = Alignment(wrap_text=True, vertical="top")  # Descripción
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
