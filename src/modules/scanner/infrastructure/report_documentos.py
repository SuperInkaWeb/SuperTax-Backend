"""
Genera un Excel (.xlsx) de los documentos/registros del Scanner a partir de filas
ya aplanadas por el frontend (una fila por documento, o por registro en las
planillas multi-registro). Reusa openpyxl, que ya es dependencia del proyecto.

Dos modos:
  - todo junto: una sola hoja con todas las filas.
  - por documento: una hoja por archivo de origen (útil en planillas).
"""
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="FF0E4F4A")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HOJA_INVALIDO = re.compile(r"[\[\]:*?/\\]")


def _valor(v: object):
    """Deja números como números (para que Excel los sume); el resto como texto."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Sí" if v else "No"
    if isinstance(v, (int, float)):
        return v
    return str(v)


def _escribir_hoja(ws, filas: list[dict], columnas: list[str], labels: dict[str, str]) -> None:
    encabezados = ["Archivo"] + [labels.get(c, c) for c in columnas]
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for fila in filas:
        ws.append([str(fila.get("archivo", ""))] + [_valor(fila.get(c)) for c in columnas])

    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for i in range(1, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def _nombre_hoja(archivo: str, usados: set[str]) -> str:
    """Nombre de hoja válido para Excel (≤31 chars, sin caracteres prohibidos, único)."""
    base = _HOJA_INVALIDO.sub(" ", str(archivo)).strip()[:28] or "Documento"
    nombre, n = base, 2
    while nombre.lower() in usados:
        sufijo = f" {n}"
        nombre = base[: 28 - len(sufijo)] + sufijo
        n += 1
    usados.add(nombre.lower())
    return nombre


def generar_excel(
    filas: list[dict],
    columnas: list[str],
    labels: dict[str, str],
    por_documento: bool = False,
) -> bytes:
    """Arma el .xlsx. Cabecera en negrita, filtro y primera fila congelada.
    Si `por_documento`, agrupa las filas en una hoja por archivo de origen."""
    wb = Workbook()

    if not por_documento:
        wb.active.title = "Documentos"
        _escribir_hoja(wb.active, filas, columnas, labels)
    else:
        # Agrupar por archivo preservando el orden de aparición.
        grupos: dict[str, list[dict]] = {}
        for fila in filas:
            grupos.setdefault(str(fila.get("archivo", "Documento")), []).append(fila)
        wb.remove(wb.active)
        usados: set[str] = set()
        for archivo, rows in grupos.items():
            _escribir_hoja(wb.create_sheet(_nombre_hoja(archivo, usados)), rows, columnas, labels)
        if not wb.sheetnames:
            _escribir_hoja(wb.create_sheet("Documentos"), [], columnas, labels)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
